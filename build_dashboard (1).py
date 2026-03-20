from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


CANONICAL_COLUMNS = {
    "client id": "Client ID",
    "loc": "Current LOC",
    "total_days": "Total Days",
    "ed_visits": "ED Visits",
    "inpt_admissions": "Inpatient Admissions",
    "mean_daily_risk": "Mean Daily Risk",
    "max_daily_risk": "Max Daily Risk",
    "days_med_mgmt_zero": "Days Med Mgmt = 0",
    "max_consec_med_miss": "Max Consecutive Med Miss",
    "mean_bprs": "Mean BPRS",
    "allocation": "Allocation",
    "annual_alloc": "Annual Allocation",
    "episode_days": "Episode Days",
    "exceeds": "Exceeds Allocation",
    "excess": "Excess Days",
    "recommendation": "Recommendation",
}

PREFERRED_RECOMMENDATION_ORDER = [
    "Upgrade Step Inpt 5 → Step Inpt 6",
    "Upgrade Step 3 → Step 4",
    "Upgrade Step 4 → Step Inpatient 5",
    "Within allocation",
    "Highest step — escalate to clinical review",
]

THIN_GRAY = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)

COLORS = {
    "navy": "1F4E78",
    "teal": "00B0F0",
    "teal_light": "D9F2F8",
    "green": "E2F0D9",
    "amber": "FFF2CC",
    "rose": "FCE4D6",
    "gray": "F3F6F9",
    "dark": "1F2937",
    "white": "FFFFFF",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the patient recommendation dashboard workbook.")
    parser.add_argument("--input", required=True, help="Path to the input CSV file.")
    parser.add_argument(
        "--output",
        required=True,
        help="Path to the output XLSX workbook.",
    )
    return parser.parse_args()


def normalize_boolean(value) -> bool:
    if isinstance(value, bool):
        return value
    if pd.isna(value):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "t"}


def load_input_csv(path: str | Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    normalized = {col.strip(): col.strip().lower() for col in df.columns}
    missing = [src for src in CANONICAL_COLUMNS if src not in normalized.values()]
    if missing:
        raise ValueError(f"Input CSV is missing required columns: {missing}")

    rename_map = {orig: CANONICAL_COLUMNS[low] for orig, low in normalized.items() if low in CANONICAL_COLUMNS}
    df = df.rename(columns=rename_map)
    df = df[[CANONICAL_COLUMNS[k] for k in CANONICAL_COLUMNS]]

    df["Exceeds Allocation"] = df["Exceeds Allocation"].map(normalize_boolean)
    df["Priority Score"] = df.apply(
        lambda row: row["Excess Days"] * (1 + row["Mean Daily Risk"] + row["Mean BPRS"])
        if row["Exceeds Allocation"]
        else 0,
        axis=1,
    )
    df["Priority Rank"] = (
        df["Priority Score"].rank(method="first", ascending=False).astype(int)
    )
    df = df.sort_values(["Priority Score", "Excess Days"], ascending=[False, False]).reset_index(drop=True)
    return df


def apply_sheet_header_style(ws, row: int, start_col: int, end_col: int) -> None:
    fill = PatternFill("solid", fgColor=COLORS["navy"])
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def style_table_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = BORDER
            if cell.row != min_row:
                cell.alignment = Alignment(vertical="center")


def autofit_widths(ws, width_overrides: dict[str, float] | None = None) -> None:
    width_overrides = width_overrides or {}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            if "\n" in value:
                value = max(value.split("\n"), key=len)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = width_overrides.get(letter, min(max_len + 2, 42))


def build_raw_data_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "Raw_Data"

    headers = list(df.columns)
    ws.append(headers)
    for _, row in df.iterrows():
        ws.append(list(row.values))

    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="tblRawData", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"

    apply_sheet_header_style(ws, 1, 1, ws.max_column)
    style_table_range(ws, 1, ws.max_row, 1, ws.max_column)

    pct_columns = ["Mean Daily Risk", "Max Daily Risk"]
    one_decimal_columns = ["Mean BPRS"]
    integer_columns = [
        "Client ID",
        "Current LOC",
        "Total Days",
        "ED Visits",
        "Inpatient Admissions",
        "Days Med Mgmt = 0",
        "Max Consecutive Med Miss",
        "Allocation",
        "Annual Allocation",
        "Episode Days",
        "Excess Days",
        "Priority Rank",
    ]

    header_index = {name: idx + 1 for idx, name in enumerate(headers)}
    for col_name in pct_columns:
        col = header_index[col_name]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col).number_format = "0.0%"
    for col_name in one_decimal_columns:
        col = header_index[col_name]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col).number_format = "0.0"
    for col_name in integer_columns:
        col = header_index[col_name]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col).number_format = "#,##0"

    priority_col = header_index["Priority Score"]
    for r in range(2, ws.max_row + 1):
        ws.cell(r, priority_col).number_format = "0.00"

    autofit_widths(
        ws,
        width_overrides={
            "P": 34,
            "Q": 14,
        },
    )


def recommendation_order(values: Iterable[str]) -> list[str]:
    seen = []
    for item in values:
        if item not in seen:
            seen.append(item)

    ordered = [item for item in PREFERRED_RECOMMENDATION_ORDER if item in seen]
    ordered.extend(item for item in seen if item not in ordered)
    return ordered


def build_summary_sheet(wb: Workbook, df: pd.DataFrame) -> dict[str, int]:
    ws = wb.create_sheet("Summary_Data")

    metrics = [
        ("Total Patients", int(df["Client ID"].count())),
        ("Patients Needing Change", int(df["Exceeds Allocation"].sum())),
        ("Within Allocation", int((~df["Exceeds Allocation"]).sum())),
        ("% Needing Change", float(df["Exceeds Allocation"].mean()) if len(df) else 0),
        ("Total Excess Days", int(df["Excess Days"].sum())),
        (
            "Avg Excess Days (Exceeded Only)",
            float(df.loc[df["Exceeds Allocation"], "Excess Days"].mean()) if df["Exceeds Allocation"].any() else 0,
        ),
        ("Avg Mean Daily Risk", float(df["Mean Daily Risk"].mean()) if len(df) else 0),
        ("Avg Mean BPRS", float(df["Mean BPRS"].mean()) if len(df) else 0),
        ("Highest Excess Days", int(df["Excess Days"].max()) if len(df) else 0),
    ]

    ws.append(["Metric", "Value"])
    for item in metrics:
        ws.append(list(item))

    ws.append([])
    ws.append([])

    rec_header_row = ws.max_row + 1
    ws.append(["Recommendation", "Patients", "Total Excess Days", "Chart Label"])
    rec_counts = df.groupby("Recommendation", dropna=False).agg(
        Patients=("Client ID", "count"),
        Total_Excess_Days=("Excess Days", "sum"),
    )
    rec_order = recommendation_order(df["Recommendation"].fillna(""))
    short_labels = {
        "Upgrade Step Inpt 5 → Step Inpt 6": "Inpt 5 → 6",
        "Upgrade Step 3 → Step 4": "Step 3 → 4",
        "Upgrade Step 4 → Step Inpatient 5": "Step 4 → Inpt 5",
        "Within allocation": "Within allocation",
        "Highest step — escalate to clinical review": "Highest step / review",
    }
    for rec in rec_order:
        row = rec_counts.loc[rec]
        ws.append([rec, int(row["Patients"]), int(row["Total_Excess_Days"]), short_labels.get(rec, rec)])
    rec_data_start = rec_header_row + 1
    rec_data_end = ws.max_row

    ws.append([])
    ws.append([])
    ws.append([])

    loc_header_row = ws.max_row + 1
    ws.append(["Current LOC", "Patients", "Avg Excess Days"])
    loc_summary = (
        df.groupby("Current LOC", dropna=False)
        .agg(Patients=("Client ID", "count"), Avg_Excess_Days=("Excess Days", "mean"))
        .reset_index()
        .sort_values("Current LOC")
    )
    for _, row in loc_summary.iterrows():
        ws.append([int(row["Current LOC"]), int(row["Patients"]), float(row["Avg_Excess_Days"])])
    loc_data_start = loc_header_row + 1
    loc_data_end = ws.max_row

    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])

    top_header_row = ws.max_row + 1
    ws.append(["Rank", "Client ID", "Current LOC", "Excess Days", "Mean Daily Risk", "Recommendation", "Priority Score"])
    top_10 = df.nsmallest(10, "Priority Rank")[
        ["Priority Rank", "Client ID", "Current LOC", "Excess Days", "Mean Daily Risk", "Recommendation", "Priority Score"]
    ]
    for _, row in top_10.iterrows():
        ws.append(list(row.values))
    top_data_start = top_header_row + 1
    top_data_end = ws.max_row

    apply_sheet_header_style(ws, 1, 1, 2)
    apply_sheet_header_style(ws, rec_header_row, 1, 4)
    apply_sheet_header_style(ws, loc_header_row, 1, 3)
    apply_sheet_header_style(ws, top_header_row, 1, 7)

    style_table_range(ws, 1, ws.max_row, 1, min(ws.max_column, 7))

    for r in range(2, 11):
        label = ws.cell(r, 1).value
        if label in {"% Needing Change", "Avg Mean Daily Risk"}:
            ws.cell(r, 2).number_format = "0.0%"
        elif label in {"Avg Excess Days (Exceeded Only)", "Avg Mean BPRS"}:
            ws.cell(r, 2).number_format = "0.00"
        else:
            ws.cell(r, 2).number_format = "#,##0"

    for r in range(rec_data_start, rec_data_end + 1):
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3).number_format = "#,##0"

    for r in range(loc_data_start, loc_data_end + 1):
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3).number_format = "0.00"

    for r in range(top_data_start, top_data_end + 1):
        ws.cell(r, 1).number_format = "#,##0"
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3).number_format = "#,##0"
        ws.cell(r, 4).number_format = "#,##0"
        ws.cell(r, 5).number_format = "0.0%"
        ws.cell(r, 7).number_format = "0.00"

    autofit_widths(ws, width_overrides={"A": 38, "D": 22, "F": 34})

    return {
        "rec_data_start": rec_data_start,
        "rec_data_end": rec_data_end,
        "loc_data_start": loc_data_start,
        "loc_data_end": loc_data_end,
        "top_data_start": top_data_start,
        "top_data_end": top_data_end,
    }


def add_kpi_card(ws, title_cell: str, value_cell: str, fill_color: str) -> None:
    title = ws[title_cell]
    value = ws[value_cell]
    title.fill = PatternFill("solid", fgColor=fill_color)
    value.fill = PatternFill("solid", fgColor=fill_color)
    title.font = Font(bold=True, color=COLORS["dark"])
    value.font = Font(bold=True, size=18, color=COLORS["dark"])
    title.alignment = Alignment(horizontal="center", vertical="center")
    value.alignment = Alignment(horizontal="center", vertical="center")
    title.border = BORDER
    value.border = BORDER


def build_dashboard_sheet(wb: Workbook, layout: dict[str, int]) -> None:
    ws = wb.create_sheet("Dashboard", 0)

    ws.merge_cells("A1:N2")
    ws["A1"] = "Patient Recommendation Dashboard"
    ws["A1"].font = Font(size=18, bold=True, color=COLORS["white"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:N3")
    ws["A3"] = "Overview of recommended patient changes based on allocation, utilization, and clinical indicators"
    ws["A3"].font = Font(italic=True, color=COLORS["dark"])
    ws["A3"].alignment = Alignment(horizontal="center")

    kpi_map = {
        "B5": ("Total Patients", "=Summary_Data!B2", COLORS["teal_light"]),
        "F5": ("Patients Needing Change", "=Summary_Data!B3", COLORS["rose"]),
        "J5": ("Within Allocation", "=Summary_Data!B4", COLORS["green"]),
        "B10": ("Total Excess Days", "=Summary_Data!B6", COLORS["amber"]),
        "F10": ("Avg Excess Days", "=Summary_Data!B7", COLORS["gray"]),
        "J10": ("Avg Mean Daily Risk", "=Summary_Data!B8", COLORS["teal_light"]),
    }
    for title_cell, (title, formula, fill) in kpi_map.items():
        value_cell = title_cell[0] + str(int(title_cell[1:]) + 1)
        ws[title_cell] = title
        ws[value_cell] = formula
        add_kpi_card(ws, title_cell, value_cell, fill)

    ws["J11"].number_format = "0.0%"
    ws["F11"].number_format = "0.00"

    ws["A15"] = "Current LOC Distribution"
    ws["H15"] = "Recommendation Mix"
    for cell_ref in ["A15", "H15", "A29"]:
        ws[cell_ref].font = Font(size=12, bold=True, color=COLORS["navy"])

    doughnut = DoughnutChart()
    doughnut.title = "Patients by Current LOC"
    doughnut.holeSize = 55
    doughnut.height = 7.5
    doughnut.width = 7.5
    doughnut.add_data(Reference(wb["Summary_Data"], min_col=2, min_row=layout["loc_data_start"], max_row=layout["loc_data_end"]), titles_from_data=False)
    doughnut.set_categories(Reference(wb["Summary_Data"], min_col=1, min_row=layout["loc_data_start"], max_row=layout["loc_data_end"]))
    ws.add_chart(doughnut, "A16")

    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = "Patients by Recommendation"
    bar.y_axis.title = "Recommendation"
    bar.x_axis.title = "Patients"
    bar.height = 7.5
    bar.width = 7.5
    bar.add_data(Reference(wb["Summary_Data"], min_col=2, min_row=layout["rec_data_start"], max_row=layout["rec_data_end"]), titles_from_data=False)
    bar.set_categories(Reference(wb["Summary_Data"], min_col=4, min_row=layout["rec_data_start"], max_row=layout["rec_data_end"]))
    ws.add_chart(bar, "H16")

    ws["A29"] = "Top 10 Highest-Priority Patients"
    dashboard_headers = ["Rank", "Client ID", "Current LOC", "Excess Days", "Mean Daily Risk", "Recommendation", "Priority Score"]
    for col_idx, header in enumerate(dashboard_headers, start=1):
        ws.cell(row=30, column=col_idx, value=header)
    apply_sheet_header_style(ws, 30, 1, 7)

    for dash_row, summary_row in enumerate(range(32, 42), start=31):
        for col_idx, summary_col in enumerate(range(1, 8), start=1):
            ws.cell(dash_row, col_idx, f"=Summary_Data!{get_column_letter(summary_col)}{summary_row}")

    for row in range(31, 41):
        ws.cell(row, 5).number_format = "0.0%"
        ws.cell(row, 7).number_format = "0.00"

    style_table_range(ws, 30, 40, 1, 7)

    ws["I30"] = (
        "How to use this dashboard\n\n"
        "• Focus first on patients with the highest priority score.\n"
        "• Recommendation mix shows where the biggest step-up demand is concentrated.\n"
        "• Current LOC distribution shows how the current panel is spread across Steps 3–6.\n\n"
        "Priority score = Excess Days × (1 + Mean Daily Risk + Mean BPRS)\n"
        "This score is a sorting aid for review, not a clinical decision rule."
    )
    ws["I30"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["I30"].fill = PatternFill("solid", fgColor=COLORS["gray"])
    ws["I30"].border = BORDER
    ws["I30"].font = Font(color=COLORS["dark"])

    for row in range(1, 41):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[30].height = 42

    for cell_range in [
        "B5:D5", "B6:D8",
        "F5:H5", "F6:H8",
        "J5:L5", "J6:L8",
        "B10:D10", "B11:D13",
        "F10:H10", "F11:H13",
        "J10:L10", "J11:L13",
        "A15:F15", "H15:N15", "A29:N29", "I30:N40",
    ]:
        ws.merge_cells(cell_range)

    # Re-populate merged KPI cells after merge.
    for title_cell, (title, formula, fill) in kpi_map.items():
        value_cell = title_cell[0] + str(int(title_cell[1:]) + 1)
        ws[title_cell] = title
        ws[value_cell] = formula
        add_kpi_card(ws, title_cell, value_cell, fill)
    ws["J11"].number_format = "0.0%"
    ws["F11"].number_format = "0.00"

    widths = {
        "A": 10, "B": 12, "C": 12, "D": 12, "E": 16, "F": 32, "G": 14,
        "H": 4, "I": 16, "J": 16, "K": 16, "L": 16, "M": 16, "N": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.sheet_view.showGridLines = False


def build_workbook(input_csv: str | Path, output_xlsx: str | Path) -> None:
    df = load_input_csv(input_csv)
    wb = Workbook()
    build_raw_data_sheet(wb, df)
    layout = build_summary_sheet(wb, df)
    build_dashboard_sheet(wb, layout)
    output_path = Path(output_xlsx)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    args = parse_args()
    build_workbook(args.input, args.output)
