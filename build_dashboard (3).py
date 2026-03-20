from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


CANONICAL_COLUMNS = {
    "client id": "Client ID",
    "total_ed": "Total ED",
    "first_ed": "First ED",
    "last_ed": "Last ED",
    "loc": "Current LOC",
    "episode_days": "Episode Days",
    "allocation": "Allocation",
    "annual_alloc": "Annual Allocation",
    "exceeds": "Exceeds Allocation",
    "excess": "Excess",
    "recommendation": "Recommendation",
}

PREFERRED_RECOMMENDATION_ORDER = [
    "Upgrade Step Inpt 5 → Step Inpt 6",
    "Upgrade Step 3 → Step 4",
    "Upgrade Step 4 → Step Inpatient 5",
    "Within allocation",
    "Highest step — escalate to clinical review",
]

THIN_GRAY = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)

COLORS = {
    "navy": "1F4E78",
    "teal": "00B0F0",
    "teal_light": "D9F2F8",
    "green": "E2F0D9",
    "amber": "FFF2CC",
    "rose": "FCE4D6",
    "gray": "F3F6F9",
    "dark": "1F2937",
    "white": "FFFFFF",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the stepped care tracking dashboard workbook.")
    parser.add_argument("--input", required=True, help="Path to the input CSV file.")
    parser.add_argument("--output", required=True, help="Path to the output XLSX workbook.")
    return parser.parse_args()


def normalize_boolean(value) -> bool:
    if isinstance(value, bool):
        return value
    if pd.isna(value):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "t"}


def recommendation_order(values: list[str]) -> list[str]:
    seen = []
    for item in values:
        if item not in seen:
            seen.append(item)
    ordered = [item for item in PREFERRED_RECOMMENDATION_ORDER if item in seen]
    ordered.extend(item for item in seen if item not in ordered)
    return ordered


def apply_sheet_header_style(ws, row: int, start_col: int, end_col: int) -> None:
    fill = PatternFill("solid", fgColor=COLORS["navy"])
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def style_table_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = BORDER
            if cell.row != min_row:
                cell.alignment = Alignment(vertical="center")


def autofit_widths(ws, width_overrides: dict[str, float] | None = None) -> None:
    width_overrides = width_overrides or {}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            if "\n" in value:
                value = max(value.split("\n"), key=len)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = width_overrides.get(letter, min(max_len + 2, 42))


def load_input_csv(path: str | Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    normalized = {col.strip(): col.strip().lower() for col in df.columns}
    missing = [src for src in CANONICAL_COLUMNS if src not in normalized.values()]
    if missing:
        raise ValueError(f"Input CSV is missing required columns: {missing}")

    rename_map = {orig: CANONICAL_COLUMNS[low] for orig, low in normalized.items() if low in CANONICAL_COLUMNS}
    df = df.rename(columns=rename_map)
    df = df[[CANONICAL_COLUMNS[k] for k in CANONICAL_COLUMNS]]

    df["Exceeds Allocation"] = df["Exceeds Allocation"].map(normalize_boolean)
    df["First ED"] = pd.to_datetime(df["First ED"], errors="coerce")
    df["Last ED"] = pd.to_datetime(df["Last ED"], errors="coerce")
    df["Current Step"] = df["Current LOC"].apply(lambda x: f"Step {int(x)}" if pd.notna(x) else "Unknown")

    ranking = (
        df.assign(
            _flag=df["Exceeds Allocation"].astype(int),
            _excess=df["Excess"].fillna(0),
            _total_ed=df["Total ED"].fillna(0),
            _episode=df["Episode Days"].fillna(0),
        )
        .sort_values(
            ["_flag", "_excess", "_total_ed", "_episode", "Client ID"],
            ascending=[False, False, False, False, True],
        )
        .reset_index(drop=True)
    )
    ranking["Priority Rank"] = range(1, len(ranking) + 1)
    df = ranking.drop(columns=["_flag", "_excess", "_total_ed", "_episode"])

    raw_cols = [
        "Client ID",
        "Total ED",
        "First ED",
        "Last ED",
        "Current LOC",
        "Current Step",
        "Episode Days",
        "Allocation",
        "Annual Allocation",
        "Exceeds Allocation",
        "Excess",
        "Recommendation",
        "Priority Rank",
    ]
    return df[raw_cols]


def build_raw_data_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "Raw_Data"

    headers = list(df.columns)
    ws.append(headers)
    for _, row in df.iterrows():
        ws.append(list(row.values))

    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="tblSteppedCareRaw", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"

    apply_sheet_header_style(ws, 1, 1, ws.max_column)
    style_table_range(ws, 1, ws.max_row, 1, ws.max_column)

    date_columns = ["First ED", "Last ED"]
    integer_columns = [
        "Client ID",
        "Total ED",
        "Current LOC",
        "Episode Days",
        "Allocation",
        "Annual Allocation",
        "Excess",
        "Priority Rank",
    ]
    header_index = {name: idx + 1 for idx, name in enumerate(headers)}

    for col_name in date_columns:
        col = header_index[col_name]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col).number_format = "yyyy-mm-dd"

    for col_name in integer_columns:
        col = header_index[col_name]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col).number_format = "#,##0"

    autofit_widths(
        ws,
        width_overrides={
            "F": 12,
            "L": 34,
        },
    )


def build_summary_sheet(wb: Workbook, df: pd.DataFrame) -> dict[str, int]:
    ws = wb.create_sheet("Summary_Data")

    exceeded_only = df.loc[df["Exceeds Allocation"]]
    metrics = [
        ("Total Clients", int(df["Client ID"].count())),
        ("Clients Exceeding Allocation", int(df["Exceeds Allocation"].sum())),
        ("Within Allocation", int((~df["Exceeds Allocation"]).sum())),
        ("% Exceeding Allocation", float(df["Exceeds Allocation"].mean()) if len(df) else 0),
        ("Total ED Visits", int(df["Total ED"].sum()) if len(df) else 0),
        ("Total Excess", int(df["Excess"].sum()) if len(df) else 0),
        ("Avg Excess (Exceeded Only)", float(exceeded_only["Excess"].mean()) if len(exceeded_only) else 0),
        ("Avg Episode Days", float(df["Episode Days"].mean()) if len(df) else 0),
        ("Earliest First ED", df["First ED"].min()),
        ("Latest Last ED", df["Last ED"].max()),
    ]

    ws.append(["Metric", "Value"])
    for item in metrics:
        ws.append(list(item))

    ws.append([])
    ws.append([])

    rec_header_row = ws.max_row + 1
    ws.append(["Recommendation", "Clients", "Total Excess", "Avg Total ED", "Chart Label"])
    rec_counts = df.groupby("Recommendation", dropna=False).agg(
        Clients=("Client ID", "count"),
        Total_Excess=("Excess", "sum"),
        Avg_Total_ED=("Total ED", "mean"),
    )
    short_labels = {
        "Upgrade Step Inpt 5 → Step Inpt 6": "Inpt 5 → 6",
        "Upgrade Step 3 → Step 4": "Step 3 → 4",
        "Upgrade Step 4 → Step Inpatient 5": "Step 4 → Inpt 5",
        "Within allocation": "Within allocation",
        "Highest step — escalate to clinical review": "Highest step / review",
    }
    for rec in recommendation_order(df["Recommendation"].fillna("").tolist()):
        row = rec_counts.loc[rec]
        ws.append([rec, int(row["Clients"]), int(row["Total_Excess"]), float(row["Avg_Total_ED"]), short_labels.get(rec, rec)])
    rec_data_start = rec_header_row + 1
    rec_data_end = ws.max_row

    ws.append([])
    ws.append([])
    ws.append([])

    loc_header_row = ws.max_row + 1
    ws.append(["Current Step", "Clients", "Exceeded", "Total Excess", "Avg Total ED"])
    loc_summary = (
        df.groupby(["Current LOC", "Current Step"], dropna=False)
        .agg(
            Clients=("Client ID", "count"),
            Exceeded=("Exceeds Allocation", "sum"),
            Total_Excess=("Excess", "sum"),
            Avg_Total_ED=("Total ED", "mean"),
        )
        .reset_index()
        .sort_values("Current LOC")
    )
    for _, row in loc_summary.iterrows():
        ws.append([
            row["Current Step"],
            int(row["Clients"]),
            int(row["Exceeded"]),
            int(row["Total_Excess"]),
            float(row["Avg_Total_ED"]),
        ])
    loc_data_start = loc_header_row + 1
    loc_data_end = ws.max_row

    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])

    top_header_row = ws.max_row + 1
    ws.append(["Rank", "Client ID", "Current Step", "Excess", "Total ED", "Episode Days", "Recommendation"])
    top_10 = df.head(10)[["Priority Rank", "Client ID", "Current Step", "Excess", "Total ED", "Episode Days", "Recommendation"]]
    for _, row in top_10.iterrows():
        ws.append(list(row.values))
    top_data_start = top_header_row + 1
    top_data_end = ws.max_row

    apply_sheet_header_style(ws, 1, 1, 2)
    apply_sheet_header_style(ws, rec_header_row, 1, 5)
    apply_sheet_header_style(ws, loc_header_row, 1, 5)
    apply_sheet_header_style(ws, top_header_row, 1, 7)

    style_table_range(ws, 1, ws.max_row, 1, min(ws.max_column, 7))

    for r in range(2, 11):
        label = ws.cell(r, 1).value
        if label == "% Exceeding Allocation":
            ws.cell(r, 2).number_format = "0.0%"
        elif label in {"Avg Excess (Exceeded Only)", "Avg Episode Days"}:
            ws.cell(r, 2).number_format = "0.0"
        elif label in {"Earliest First ED", "Latest Last ED"}:
            ws.cell(r, 2).number_format = "yyyy-mm-dd"
        else:
            ws.cell(r, 2).number_format = "#,##0"

    for r in range(rec_data_start, rec_data_end + 1):
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3).number_format = "#,##0"
        ws.cell(r, 4).number_format = "0.0"

    for r in range(loc_data_start, loc_data_end + 1):
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3).number_format = "#,##0"
        ws.cell(r, 4).number_format = "#,##0"
        ws.cell(r, 5).number_format = "0.0"

    for r in range(top_data_start, top_data_end + 1):
        for c in [1, 2, 4, 5, 6]:
            ws.cell(r, c).number_format = "#,##0"

    autofit_widths(ws, width_overrides={"A": 34, "E": 18, "G": 34})

    return {
        "rec_data_start": rec_data_start,
        "rec_data_end": rec_data_end,
        "loc_data_start": loc_data_start,
        "loc_data_end": loc_data_end,
        "top_data_start": top_data_start,
        "top_data_end": top_data_end,
    }


def add_kpi_card(ws, title_cell: str, value_cell: str, fill_color: str) -> None:
    title = ws[title_cell]
    value = ws[value_cell]
    title.fill = PatternFill("solid", fgColor=fill_color)
    value.fill = PatternFill("solid", fgColor=fill_color)
    title.font = Font(bold=True, color=COLORS["dark"])
    value.font = Font(bold=True, size=18, color=COLORS["dark"])
    title.alignment = Alignment(horizontal="center", vertical="center")
    value.alignment = Alignment(horizontal="center", vertical="center")
    title.border = BORDER
    value.border = BORDER


def build_dashboard_sheet(wb: Workbook, layout: dict[str, int]) -> None:
    ws = wb.create_sheet("Dashboard", 0)

    ws.merge_cells("A1:N2")
    ws["A1"] = "Stepped Care Tracking Dashboard"
    ws["A1"].font = Font(size=18, bold=True, color=COLORS["white"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:N3")
    ws["A3"] = "Daily view of clients exceeding stepped-care allocation and recommended level changes"
    ws["A3"].font = Font(italic=True, color=COLORS["dark"])
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:N4")
    ws["A4"] = f"Last refreshed: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A4"].font = Font(size=10, color=COLORS["dark"])
    ws["A4"].alignment = Alignment(horizontal="center")

    kpi_map = {
        "B6": ("Total Clients", "=Summary_Data!B2", COLORS["teal_light"]),
        "F6": ("Exceeding Allocation", "=Summary_Data!B3", COLORS["rose"]),
        "J6": ("Within Allocation", "=Summary_Data!B4", COLORS["green"]),
        "B11": ("Total Excess", "=Summary_Data!B7", COLORS["amber"]),
        "F11": ("Total ED Visits", "=Summary_Data!B6", COLORS["gray"]),
        "J11": ("Avg Excess (Exceeded)", "=Summary_Data!B8", COLORS["teal_light"]),
    }
    for title_cell, (title, formula, fill) in kpi_map.items():
        value_cell = title_cell[0] + str(int(title_cell[1:]) + 1)
        ws[title_cell] = title
        ws[value_cell] = formula
        add_kpi_card(ws, title_cell, value_cell, fill)

    ws["J12"].number_format = "0.0"

    ws["A16"] = "Current Step Distribution"
    ws["H16"] = "Recommendation Mix"
    for cell_ref in ["A16", "H16", "A30"]:
        ws[cell_ref].font = Font(size=12, bold=True, color=COLORS["navy"])

    doughnut = DoughnutChart()
    doughnut.title = "Clients by Current Step"
    doughnut.holeSize = 55
    doughnut.height = 7.5
    doughnut.width = 7.5
    doughnut.add_data(Reference(wb["Summary_Data"], min_col=2, min_row=layout["loc_data_start"], max_row=layout["loc_data_end"]), titles_from_data=False)
    doughnut.set_categories(Reference(wb["Summary_Data"], min_col=1, min_row=layout["loc_data_start"], max_row=layout["loc_data_end"]))
    ws.add_chart(doughnut, "A17")

    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = "Clients by Recommendation"
    bar.y_axis.title = "Recommendation"
    bar.x_axis.title = "Clients"
    bar.height = 7.5
    bar.width = 7.5
    bar.add_data(Reference(wb["Summary_Data"], min_col=2, min_row=layout["rec_data_start"], max_row=layout["rec_data_end"]), titles_from_data=False)
    bar.set_categories(Reference(wb["Summary_Data"], min_col=5, min_row=layout["rec_data_start"], max_row=layout["rec_data_end"]))
    ws.add_chart(bar, "H17")

    ws["A30"] = "Top 10 Clients for Review"
    dashboard_headers = ["Rank", "Client ID", "Current Step", "Excess", "Total ED", "Episode Days", "Recommendation"]
    for col_idx, header in enumerate(dashboard_headers, start=1):
        ws.cell(row=31, column=col_idx, value=header)
    apply_sheet_header_style(ws, 31, 1, 7)

    for dash_row, summary_row in enumerate(range(layout["top_data_start"], layout["top_data_end"] + 1), start=32):
        for col_idx, summary_col in enumerate(range(1, 8), start=1):
            ws.cell(dash_row, col_idx, f"=Summary_Data!{get_column_letter(summary_col)}{summary_row}")

    style_table_range(ws, 31, 41, 1, 7)
    for row in range(32, 42):
        for c in [1, 2, 4, 5, 6]:
            ws.cell(row, c).number_format = "#,##0"

    ws["I31"] = (
        "How to use this dashboard\n\n"
        "• Start with the top review list, which is ordered by exceed flag, excess, total ED, and episode days.\n"
        "• Recommendation mix shows where stepped-care movement is concentrated.\n"
        "• Current step distribution shows how the panel is spread across Steps 3–6.\n\n"
        "This dashboard is an operational prioritization view and does not replace clinical judgment."
    )
    ws["I31"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["I31"].fill = PatternFill("solid", fgColor=COLORS["gray"])
    ws["I31"].border = BORDER
    ws["I31"].font = Font(color=COLORS["dark"])

    for row in range(1, 42):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[31].height = 42

    for cell_range in [
        "B6:D6", "B7:D9",
        "F6:H6", "F7:H9",
        "J6:L6", "J7:L9",
        "B11:D11", "B12:D14",
        "F11:H11", "F12:H14",
        "J11:L11", "J12:L14",
        "A16:F16", "H16:N16", "A30:N30", "I31:N41",
    ]:
        ws.merge_cells(cell_range)

    for title_cell, (title, formula, fill) in kpi_map.items():
        value_cell = title_cell[0] + str(int(title_cell[1:]) + 1)
        ws[title_cell] = title
        ws[value_cell] = formula
        add_kpi_card(ws, title_cell, value_cell, fill)
    ws["J12"].number_format = "0.0"

    widths = {
        "A": 10, "B": 12, "C": 14, "D": 12, "E": 12, "F": 14, "G": 34,
        "H": 4, "I": 16, "J": 16, "K": 16, "L": 16, "M": 16, "N": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.sheet_view.showGridLines = False


def build_workbook(input_csv: str | Path, output_xlsx: str | Path) -> None:
    df = load_input_csv(input_csv)
    wb = Workbook()
    build_raw_data_sheet(wb, df)
    layout = build_summary_sheet(wb, df)
    build_dashboard_sheet(wb, layout)

    output_path = Path(output_xlsx)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    args = parse_args()
    build_workbook(args.input, args.output)
